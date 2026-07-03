"""Run: python check_excel_versions.py"""
import openpyxl, os

folder = os.path.join(os.path.dirname(__file__), 'excel')
files = [
    'PL_Apr-25_to_Mar-26_V11.xlsx',
    'PL_Apr-25_to_Mar-26_V12.xlsx',
    'PL_Apr-25_to_Mar-26_V13.xlsx',
]

TALLY = {
    'Revenue':  68.00,
    'GP':       19.49,
    'Overhead':  7.20,
    'NP':       12.44,
}

for fname in files:
    path = os.path.join(folder, fname)
    if not os.path.exists(path):
        print(f"\n❌ NOT FOUND: {fname}")
        continue

    print(f"\n{'='*60}")
    print(f"FILE: {fname}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Extract all rows with values
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(v for v in row):
            name = str(row[0] or '').strip()
            total = row[-1]
            if name:
                rows.append((name, total))

    # Print full structure
    print(f"\n{'Particulars':50s} {'Total':>12s}")
    print("-"*65)
    for name, val in rows:
        if name == 'Particulars': continue
        val_str = f"{val:.4f}" if isinstance(val, (int, float)) else str(val or '')
        print(f"  {name:48s} {val_str:>12s}")

    # Check key values
    print(f"\n--- KEY VALUES vs TALLY ---")
    found = {}
    for name, val in rows:
        n = name.upper()
        if 'GROSS PROFIT' in n and 'NET' not in n: found['GP'] = val
        if 'NET PROFIT' in n: found['NP'] = val
        if 'SUB-TOTAL' in n and 'OVERHEAD' not in n and 'GP' not in found: pass
        if 'OVERHEAD' in n or 'INDIRECT EXP' in n:
            if isinstance(val, (int,float)) and abs(val) > 0: found['Overhead'] = val
        if 'REVENUE' in n or 'SALES ACCOUNTS' in n:
            if isinstance(val, (int,float)) and abs(val) > 0 and 'Revenue' not in found:
                found['Revenue'] = val

    for k, tv in TALLY.items():
        pv = found.get(k)
        if pv is not None:
            diff = float(pv) - tv
            ok = '✅' if abs(diff) < 0.05 else '❌'
            print(f"  {k:15s}: Portal={float(pv):>8.4f}  Tally={tv:>8.4f}  Diff={diff:>+8.4f}  {ok}")
        else:
            print(f"  {k:15s}: ⚠️  Not found in Excel")

