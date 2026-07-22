"""
Cash Flow Report — Indirect Method
Derived purely from pl_data (monthly P&L net values) and bs_data
(monthly closing balances). No new Tally sync, no new tables.

═══════════════════════════════════════════════════════════════
SIGN CONVENTION (read this before touching any formula below)
═══════════════════════════════════════════════════════════════
1. Raw bs_data.closing_bal sign was verified empirically (not assumed)
   by inspecting individual ledgers across all 3 synced companies:
   Cash-in-hand, Bank Accounts, Fixed Assets and Closing Stock all
   carry a NEGATIVE raw balance when the balance is "normal" for that
   asset. Sundry Creditors, Capital Account, Loans and Duties & Taxes
   all carry POSITIVE raw balances when normal for that liability.

   So: this database's convention is "credit-positive" — asset
   ledgers are stored negative, liability/equity ledgers positive.

   NORMALISED value used everywhere below:
     - asset bucket:     normalised = -1 * closing_bal   (so a normal
                          asset balance becomes a positive number)
     - liability/equity: normalised = +1 * closing_bal   (already
                          positive when normal)

2. Classification is HYBRID and matches on LEDGER_NAME, not
   tally_group/mis_group. Spot-checking raw rows showed tally_group is
   unreliable for this client's synced bs_data (e.g. ledger
   'Fixed Assets' was tagged tally_group='Profit & Loss A/c' — the
   field is shifted by one row in the Tally XML export). This matches
   dashboard.py's own approach, which already avoids tally_group.

3. Tally's BS is a NESTED hierarchy: every section stores BOTH a
   rolled-up group-total row (ledger_name == the group name, e.g.
   "Fixed Assets" = 70.5L) AND its member ledgers (Buildings,
   Computers, Plant & Machinery … which re-sum to 70.5L). Counting
   both double-counts the section — that was the original CFI bug
   (portal showed 2× the true fixed-asset movement). The fix
   (_bs_monthly + CASHFLOW_GROUP_TOTAL_NAMES) PREFERS the group-total
   row and skips members; it only sums members for a bucket whose
   group-total row is absent in that month. Parent rows
   (Current Assets / Current Liabilities) are always skipped — they
   roll up children and would double-count.

4. Cash-flow sign rules applied to every MoM Δ of a normalised balance:
     - Asset increase   → cash OUTFLOW  → contributes  -Δ
     - Liability increase → cash INFLOW → contributes  +Δ

5. EBITDA row. The spec's "EBITDA = GP + Indirect Incomes − Overhead,
   i.e. operating profit BEFORE Depreciation, Interest, Tax" is
   algebraically identical to:
       EBITDA = Net Profit + Depreciation + Interest Paid − Interest Received
   (because Net Profit already has Depreciation and Interest Paid
   deducted via Overhead, and Interest Received added via Indirect
   Incomes). To avoid double-counting, this file shows it as three
   transparent rows — Net Profit, +Depreciation, +/-Interest — whose
   SUBTOTAL is labelled "EBITDA / Operating Profit before Working
   Capital Changes". This is the standard, non-duplicating Indirect
   Method starting point.

═══════════════════════════════════════════════════════════════
VALIDATION NOTE (honest reporting, not a guess)
═══════════════════════════════════════════════════════════════
This sign convention + classification was tested against the real
synced database before writing this page. Result:
  - Unique Steel Products: ties out to within ~0.2% of the actual
    closing cash balance for most months from Apr-2025 onward. Months
    before that show a constant, clearly-erroneous P&L figure (likely
    placeholder/duplicated sync data for that stretch) — a pre-existing
    data issue, not a formula issue.
  - Intellichemie Industries LLP and Avenuecorp India Private Limited:
    do NOT reliably tie out — the gap is large and not systematic
    (sign or bucket fixes alone did not resolve it). This likely means
    those two companies' BS sync has ledger-naming or fiscal-year
    quirks not covered by the keyword lists above, which needs
    checking against live Tally data, not just SQL inspection.
The ✅/⚠️ status shown per month is this diagnostic made visible,
rather than hidden — use it to see exactly which months are trustworthy.
"""
import io
import streamlit as st

from core.db        import get_conn
from core.theme     import chart_layout, CHART_COLORS, CHART_PALETTE
from core.constants import (MONTHS, CASHFLOW_ASSET_BUCKETS,
                            CASHFLOW_GROUP_TOTAL_NAMES, CASHFLOW_PARENT_SKIP,
                            CASHFLOW_MEMBER_KEYWORDS)
from core.utils     import fmt_inr
from portal_pages.dashboard import _calc


# ── BS CLASSIFICATION (hybrid: group-total preferred, member fallback) ──
def _normalize(bucket: str, raw: float) -> float:
    return -raw if bucket in CASHFLOW_ASSET_BUCKETS else raw


# ── PER-MONTH BS AGGREGATION ─────────────────────────────────────
def _bs_monthly(bs_rows):
    """
    Hybrid classification to avoid Tally's nested double-counting
    (see CASHFLOW_GROUP_TOTAL_NAMES docstring in core.constants).

    For each month, per bucket:
      1. If a group-total row exists (ledger_name is a known group name),
         use it — this is Tally's own rolled-up figure. Skip members.
      2. Otherwise, sum the member ledgers matched by keyword.
      3. Never assign parent rows (Current Assets / Current Liabilities)
         to any bucket — they roll up children and would double-count.

    Bank OD / Overdraft handling: Bank OD A/c sits under Tally's
    "Loans (Liability)" group (confirmed by checking tally_group on the
    real synced data), so that group-total row's figure already
    INCLUDES the overdraft balance. Per business decision, overdraft is
    treated as a cash-equivalent (short-term, tied to bank operations)
    rather than a financing loan — so its balance is: (a) added to
    cash_actual like a normal bank ledger, and (b) NETTED OUT of the
    'loans_liability' bucket so it isn't counted twice (once as cash,
    once inside the Loans group-total that already rolled it up).

    Returns:
      bucket_by_month:      {(yr,mo): {bucket: normalised_total}}
      cash_actual_by_month: {(yr,mo): actual cash+bank total (positive)}
    """
    from collections import defaultdict

    rows_by_month = defaultdict(list)
    for r in bs_rows:
        rows_by_month[(int(r['year']), int(r['month']))].append(r)

    bucket_by_month = defaultdict(lambda: defaultdict(float))
    cash_actual     = defaultdict(float)

    def _is_bank_od(n: str) -> bool:
        return 'bank od' in n or 'overdraft' in n or 'bank occ' in n

    def _is_plug_entry(n: str) -> bool:
        """Exclude Tally's plug/suspense entries — not real cash movements."""
        return (
            'difference in opening' in n or
            'suspense' in n or
            'opening balance difference' in n
        )

    for key, mrows in rows_by_month.items():
        # 1. Group-total rows (authoritative)
        present_groups = {}
        bank_od_raw = 0.0  # raw (liability-sign) sum of OD ledgers this month
        for r in mrows:
            n = (r['ledger_name'] or '').lower().strip()
            if n in CASHFLOW_GROUP_TOTAL_NAMES:
                b = CASHFLOW_GROUP_TOTAL_NAMES[n]
                present_groups[b] = present_groups.get(b, 0) + \
                    _normalize(b, r['closing_bal'] or 0)
            elif _is_bank_od(n):
                bank_od_raw += (r['closing_bal'] or 0)

        # Net Bank OD out of Loans (Liability) — Tally's own group-total
        # already rolled it up, so subtract its raw (liability-sign)
        # contribution before it's used for CFF.
        if bank_od_raw and 'loans_liability' in present_groups:
            present_groups['loans_liability'] -= bank_od_raw

        # 2. Member fallback — only for buckets with no group total
        member_sums = defaultdict(float)
        for r in mrows:
            n = (r['ledger_name'] or '').lower().strip()
            if n in CASHFLOW_GROUP_TOTAL_NAMES or n in CASHFLOW_PARENT_SKIP:
                continue
            if _is_bank_od(n):
                continue  # already handled above (moved to cash)
            for bucket, kws in CASHFLOW_MEMBER_KEYWORDS.items():
                if bucket in present_groups:
                    continue  # group total already covers this bucket
                if any(kw in n for kw in kws):
                    member_sums[bucket] += _normalize(bucket, r['closing_bal'] or 0)
                    break

        for b, v in present_groups.items():
            bucket_by_month[key][b] += v
        for b, v in member_sums.items():
            bucket_by_month[key][b] += v

        # Actual cash+bank: normal cash/bank ledgers use abs() (their raw
        # is reliably negative when normal). Bank OD is DIFFERENT — its
        # raw is liability-positive when the account is overdrawn, which
        # must REDUCE cash (negate, don't abs()), otherwise an overdrawn
        # balance would be wrongly added as positive cash.
        for r in mrows:
            n   = (r['ledger_name'] or '').lower()
            raw = r['closing_bal'] or 0
            if 'bank account' in n or 'cash-in-hand' in n or 'cash in hand' in n:
                cash_actual[key] += abs(raw)
            elif _is_bank_od(n):
                cash_actual[key] += -raw

    return bucket_by_month, cash_actual


# ── PER-MONTH P&L → CFO COMPONENTS ───────────────────────────────
def _pl_monthly(pl_rows):
    from collections import defaultdict
    buckets = defaultdict(list)
    for r in pl_rows:
        buckets[(int(r['year']), int(r['month']))].append(r)
    return buckets


def _cf_components(rows):
    """
    Returns (p, depreciation, interest_paid, interest_received) for one
    month's pl_data rows, where `p` is dashboard.py's standard _calc()
    result (revenue/gp/overhead/np/etc — unchanged, reused as-is per
    the hard rule "do not modify P&L/dashboard calculation logic").
    """
    p = _calc(rows)
    dep = interest_paid = interest_received = 0.0
    for r in rows:
        ln  = str(r['ledger_name'] or '').lower()
        mg  = str(r['mis_group']   or '').lower()
        tg  = str(r['tally_group'] or '').lower()
        val = abs(r['net'] or 0)

        if 'deprecia' in ln or 'deprecia' in mg:
            dep += val
        elif tg == 'indirect expenses' and (
            'finance cost' in mg or 'interest' in mg or 'interest' in ln
        ):
            interest_paid += val
        elif tg == 'indirect incomes' and (
            'interest received' in mg or 'interest' in ln
        ):
            interest_received += val

    return p, dep, interest_paid, interest_received


# ── BUILD THE FULL STATEMENT (list of row dicts) ─────────────────
def _build_statement(pl_buckets, bs_buckets, cash_actual, months, mo_labels):
    """
    months: ordered list of (yr,mo) tuples INCLUDING one prior month
            (months[0] is the prior month, used only for deltas —
            it is not rendered as its own column).
    Returns: (table_rows, per_month dict keyed by mo_label with all
              computed line values, for chart use)
    """
    per_month = {}

    for i in range(1, len(months)):
        key      = months[i]
        prev_key = months[i - 1]
        lbl      = mo_labels[i - 1]

        rows = pl_buckets.get(key, [])
        p, dep, int_paid, int_recv = _cf_components(rows)
        net_profit = p['np']

        bs_now  = bs_buckets.get(key, {})
        bs_prev = bs_buckets.get(prev_key, {})

        def delta(bucket):
            return bs_now.get(bucket, 0) - bs_prev.get(bucket, 0)

        d_debtors  = delta('debtors')
        d_stock    = delta('stock')
        d_loanadv  = delta('loans_advances_asset')
        d_deposits = delta('deposits')
        d_oca      = delta('other_current_assets')
        wc_assets  = -(d_debtors + d_stock + d_loanadv + d_deposits + d_oca)

        d_creditors = delta('creditors')
        d_prov      = delta('provisions')
        d_duties    = delta('duties_taxes')
        d_ocl       = delta('other_current_liabilities')
        wc_liab     = d_creditors + d_prov + d_duties + d_ocl

        ebitda = net_profit + dep + int_paid - int_recv
        cfo    = ebitda + wc_assets + wc_liab

        d_fa  = delta('fixed_assets')
        d_inv = delta('investments')
        cfi   = -d_fa - d_inv

        d_cap = delta('capital')
        d_loan = delta('loans_liability')
        cff = (d_cap - net_profit) + d_loan

        net_cash = cfo + cfi + cff
        opening  = cash_actual.get(prev_key, 0)
        closing_computed = opening + net_cash
        closing_actual   = cash_actual.get(key, 0)
        diff = closing_computed - closing_actual
        tol  = max(1000.0, 0.005 * abs(closing_actual))
        tied = abs(diff) < tol

        per_month[lbl] = dict(
            net_profit=net_profit, depreciation=dep, interest_paid=int_paid,
            interest_received=int_recv, ebitda=ebitda,
            d_debtors=d_debtors, d_stock=d_stock, d_loanadv=d_loanadv,
            d_deposits=d_deposits, d_oca=d_oca, wc_assets=wc_assets,
            d_creditors=d_creditors, d_prov=d_prov, d_duties=d_duties,
            d_ocl=d_ocl, wc_liab=wc_liab, cfo=cfo,
            d_fa=d_fa, d_inv=d_inv, cfi=cfi,
            d_cap=d_cap, d_loan=d_loan, cff=cff,
            net_cash=net_cash, opening=opening,
            closing_computed=closing_computed, closing_actual=closing_actual,
            diff=diff, tied=tied,
        )

    return per_month


# ── MAIN PAGE ─────────────────────────────────────────────────────
def show_cash_flow(user):
    conn = get_conn()

    company_id   = st.session_state.get('global_company_id')
    company_name = st.session_state.get('global_company_name', '')
    from_lbl     = st.session_state.get('global_from_lbl')
    to_lbl       = st.session_state.get('global_to_lbl')
    from_yr      = st.session_state.get('global_from_yr')
    from_mo      = st.session_state.get('global_from_mo')
    to_yr        = st.session_state.get('global_to_yr')
    to_mo        = st.session_state.get('global_to_mo')

    if not company_id or not from_lbl:
        conn.close()
        st.info("Select a company and date range from the sidebar to view the cash flow report.")
        return

    if (from_yr, from_mo) > (to_yr, to_mo):
        conn.close()
        st.error("'From' cannot be after 'To'.")
        return

    # Available months (same pattern as reports.py)
    avail = conn.execute(
        "SELECT DISTINCT year, month FROM pl_data "
        "WHERE company_id=? ORDER BY year, month", (company_id,)
    ).fetchall()
    if not avail:
        conn.close()
        st.info("No data synced yet. Please sync from Tally first.")
        return
    avail_ym = [(int(r['year']), int(r['month'])) for r in avail]

    sel_months = [ym for ym in avail_ym if (from_yr, from_mo) <= ym <= (to_yr, to_mo)]
    if not sel_months:
        conn.close()
        st.info("No data in the selected range.")
        return
    mo_labels = [f"{MONTHS[m-1]}-{str(y)[2:]}" for y, m in sel_months]

    # One prior month (for the first month's MoM delta + opening cash)
    idx0 = avail_ym.index(sel_months[0])
    prior_month = [avail_ym[idx0 - 1]] if idx0 > 0 else []
    months_with_prior = prior_month + sel_months

    # ── STICKY HEADER ──────────────────────────────────────
    st.markdown(f"""
        <div class="page-sticky-header">
            <div class="page-title-row">💵 Cash Flow Statement</div>
            <div class="page-company-row">🏢 {company_name}</div>
        </div>
    """, unsafe_allow_html=True)
    st.caption(f"📅 {from_lbl} → {to_lbl} &nbsp;|&nbsp; {len(sel_months)} months "
               f"&nbsp;|&nbsp; Indirect Method", unsafe_allow_html=True)
    if not prior_month:
        st.caption("⚠️ No prior month available before the range start — "
                   "the first month's working-capital change and opening "
                   "cash could not be computed and is shown as 0.")
    st.markdown("---")

    # ── FETCH ──────────────────────────────────────────────
    yr0, mo0 = months_with_prior[0]
    pl_rows = conn.execute("""
        SELECT ledger_name, tally_group, mis_group, year, month, net
        FROM pl_data
        WHERE company_id=?
          AND ((year > ?) OR (year = ? AND month >= ?))
          AND ((year < ?) OR (year = ? AND month <= ?))
    """, (company_id, from_yr, from_yr, from_mo, to_yr, to_yr, to_mo)).fetchall()

    bs_rows = conn.execute("""
        SELECT ledger_name, year, month, closing_bal
        FROM bs_data
        WHERE company_id=?
          AND ((year > ?) OR (year = ? AND month >= ?))
          AND ((year < ?) OR (year = ? AND month <= ?))
    """, (company_id, yr0, yr0, mo0, to_yr, to_yr, to_mo)).fetchall()
    conn.close()

    if not pl_rows:
        st.info("No P&L data for selected period.")
        return

    import plotly.graph_objects as go

    pl_buckets            = _pl_monthly(pl_rows)
    bs_buckets, cash_actual = _bs_monthly(bs_rows)
    per_month = _build_statement(pl_buckets, bs_buckets, cash_actual,
                                  months_with_prior, mo_labels)

    if not per_month:
        st.info("Not enough consecutive months of data to compute cash flow.")
        return

    # ── KPI STRIP ──────────────────────────────────────────
    total_net_cash = sum(m['net_cash'] for m in per_month.values())
    total_cfo      = sum(m['cfo']      for m in per_month.values())
    total_cfi      = sum(m['cfi']      for m in per_month.values())
    total_cff      = sum(m['cff']      for m in per_month.values())
    last_lbl       = mo_labels[-1]
    last           = per_month.get(last_lbl)
    tied_count     = sum(1 for m in per_month.values() if m['tied'])
    total_count    = len(per_month)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("💰 Net Cash Generated (period)", fmt_inr(total_net_cash))
    c2.metric("⚙️ CFO (period)", fmt_inr(total_cfo))
    c3.metric("📈 CFI (period)", fmt_inr(total_cfi))
    c4.metric("💵 CFF (period)", fmt_inr(total_cff))
    if last:
        c5.metric(f"🏦 Closing Cash ({last_lbl})", fmt_inr(last['closing_actual']))
    else:
        c5.metric("🏦 Closing Cash", "—")

    # Reconciliation status shown as a caption (the per-month ✅/⚠️ flags
    # live in the statement table's Difference row).
    recon_ok = tied_count == total_count
    st.caption(
        (f"🔍 Reconciliation: {tied_count}/{total_count} months tied out "
         f"within tolerance.")
        + ("" if recon_ok else
           " ⚠️ Some months need review — see the Difference row in the "
           "statement below.")
    )

    st.markdown("---")

    tab1, tab2 = st.tabs(["📊 Chart View", "📋 Cash Flow Statement"])

    # ── TAB 1: CHART VIEW ──────────────────────────────────
    with tab1:
        cfo_l = [per_month[l]['cfo'] for l in mo_labels]
        cfi_l = [per_month[l]['cfi'] for l in mo_labels]
        cff_l = [per_month[l]['cff'] for l in mo_labels]
        close_l = [per_month[l]['closing_actual'] for l in mo_labels]
        close_computed_l = [per_month[l]['closing_computed'] for l in mo_labels]

        col1, col2 = st.columns(2)
        with col1:
            st.markdown('<div class="chart-card"><p class="chart-title">'
                        '📊 CFO vs CFI vs CFF by Month</p>', unsafe_allow_html=True)
            fig = go.Figure()
            fig.add_trace(go.Bar(x=mo_labels, y=cfo_l, name='CFO',
                                  marker=dict(color=CHART_COLORS['mint'])))
            fig.add_trace(go.Bar(x=mo_labels, y=cfi_l, name='CFI',
                                  marker=dict(color=CHART_COLORS['amber'])))
            fig.add_trace(go.Bar(x=mo_labels, y=cff_l, name='CFF',
                                  marker=dict(color=CHART_COLORS['violet'])))
            chart_layout(fig, height=320, barmode='group',
                         legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                     font=dict(color=CHART_COLORS['text'], size=10)))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
            st.markdown('</div>', unsafe_allow_html=True)

        with col2:
            st.markdown('<div class="chart-card"><p class="chart-title">'
                        '🏦 Closing Cash Balance Trend</p>', unsafe_allow_html=True)
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(
                x=mo_labels, y=close_l, name='Actual Closing Cash',
                mode='lines+markers', fill='tozeroy',
                line=dict(color=CHART_COLORS['blue'], width=2.5)))
            fig2.add_trace(go.Scatter(
                x=mo_labels, y=close_computed_l, name='Computed Closing Cash',
                mode='lines+markers', line=dict(color=CHART_COLORS['red'],
                                                 width=1.5, dash='dot')))
            chart_layout(fig2, height=320,
                         legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                     font=dict(color=CHART_COLORS['text'], size=10)))
            st.plotly_chart(fig2, use_container_width=True, config={'displayModeBar': False})
            st.markdown('</div>', unsafe_allow_html=True)

        # Waterfall for the latest month
        st.markdown('<div class="chart-card"><p class="chart-title">'
                    f'💧 Cash Bridge — {last_lbl}</p>', unsafe_allow_html=True)
        if last:
            wf_labels = ["Opening Cash", "EBITDA", "Working Capital",
                         "CFO", "CFI", "CFF", "Closing Cash (computed)"]
            wf_values = [last['opening'], last['ebitda'],
                         last['wc_assets'] + last['wc_liab'],
                         0, last['cfi'], last['cff'], last['closing_computed']]
            measures = ["absolute", "relative", "relative",
                        "total", "relative", "relative", "total"]
            fig3 = go.Figure(go.Waterfall(
                x=wf_labels, y=wf_values, measure=measures,
                increasing=dict(marker=dict(color=CHART_COLORS['mint'])),
                decreasing=dict(marker=dict(color=CHART_COLORS['red'])),
                totals=dict(marker=dict(color=CHART_COLORS['blue'])),
            ))
            chart_layout(fig3, height=340, showlegend=False)
            st.plotly_chart(fig3, use_container_width=True, config={'displayModeBar': False})
        st.markdown('</div>', unsafe_allow_html=True)

    # ── TAB 2: STATEMENT TABLE ──────────────────────────────
    with tab2:
        table_rows = _render_statement_rows(per_month, mo_labels)

        # ── EXPORT — dropdown: .xlsx or .csv, filename includes client ──
        report_base_name = f"Cash Flow — {company_name}".strip()
        for ch in '\\/:*?"<>|':
            report_base_name = report_base_name.replace(ch, '-')

        col_title, col_dl = st.columns([6, 1])
        with col_title:
            st.markdown("**📋 Cash Flow Statement — Monthly (Indirect Method)**")
        with col_dl:
            with st.popover("📥 Export", use_container_width=True):
                st.caption(report_base_name)
                excel_bytes = _generate_cashflow_excel(
                    per_month, mo_labels, from_lbl, to_lbl, company_name
                )
                st.download_button(
                    label="⬇️ Download as .xlsx",
                    data=excel_bytes,
                    file_name=f"{report_base_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="cf_dl_xlsx",
                )
                import pandas as pd
                csv_df = pd.DataFrame(table_rows)
                csv_cols = ["Particulars"] + mo_labels + ["Total"]
                csv_df = csv_df.reindex(columns=csv_cols, fill_value="")
                csv_bytes = csv_df.to_csv(index=False).encode("utf-8-sig")
                st.download_button(
                    label="⬇️ Download as .csv",
                    data=csv_bytes,
                    file_name=f"{report_base_name}.csv",
                    mime="text/csv",
                        use_container_width=True,

# ── STATEMENT ROW BUILDER (shared by on-screen table + CSV) ──────
def _render_statement_rows(per_month, mo_labels):
    def row(label, key, rtype='ledger', indent=0, is_pct=False, sign=1):
        r = {'Particulars': ('    ' * indent) + label, '_type': rtype}
        total = 0.0
        for lbl in mo_labels:
            v = per_month.get(lbl, {}).get(key, 0) * sign
            total += v
            r[lbl] = fmt_inr(v) if v else "—"
        # Total column — simple sum of all months (per user's choice)
        r['Total'] = fmt_inr(total) if total else "—"
        return r

    rows = []
    rows.append({'Particulars': 'I. CASH FLOW FROM OPERATING ACTIVITY (CFO)', '_type': 'header'})
    rows.append(row('Net Profit (as per P&L)', 'net_profit', indent=1))
    rows.append(row('Add: Depreciation', 'depreciation', indent=1))
    rows.append(_interest_row(per_month, mo_labels))
    rows.append(row('= EBITDA / Operating Profit before Working Capital Changes',
                    'ebitda', rtype='subtotal', indent=1))

    rows.append({'Particulars': '    Working Capital Changes', '_type': 'ledger'})
    rows.append(row('(Increase)/Decrease in Debtors', 'd_debtors', indent=2, sign=-1))
    rows.append(row('(Increase)/Decrease in Stock', 'd_stock', indent=2, sign=-1))
    rows.append(row('(Increase)/Decrease in Loans & Advances', 'd_loanadv', indent=2, sign=-1))
    rows.append(row('(Increase)/Decrease in Deposits', 'd_deposits', indent=2, sign=-1))
    rows.append(row('(Increase)/Decrease in Other Current Assets', 'd_oca', indent=2, sign=-1))
    rows.append(row('= Net (Decrease)/Increase in Current Assets', 'wc_assets',
                    rtype='subtotal', indent=2))
    rows.append(row('Increase/(Decrease) in Creditors', 'd_creditors', indent=2))
    rows.append(row('Increase/(Decrease) in Provisions', 'd_prov', indent=2))
    rows.append(row('Increase/(Decrease) in Duties & Taxes', 'd_duties', indent=2))
    rows.append(row('Increase/(Decrease) in Other Current Liabilities', 'd_ocl', indent=2))
    rows.append(row('= Net Increase/(Decrease) in Current Liabilities', 'wc_liab',
                    rtype='subtotal', indent=2))
    rows.append(row('= CASH FLOW FROM OPERATING ACTIVITY', 'cfo', rtype='cfo'))

    rows.append({'Particulars': '', '_type': 'spacer'})
    rows.append({'Particulars': 'II. CASH FLOW FROM INVESTING ACTIVITY (CFI)', '_type': 'header'})
    rows.append(row('(Increase)/Decrease in Fixed Assets', 'd_fa', indent=1, sign=-1))
    rows.append(row('(Increase)/Decrease in Investments', 'd_inv', indent=1, sign=-1))
    rows.append(row('= CASH FLOW FROM INVESTING ACTIVITY', 'cfi', rtype='cfi'))

    rows.append({'Particulars': '', '_type': 'spacer'})
    rows.append({'Particulars': 'III. CASH FLOW FROM FINANCING ACTIVITY (CFF)', '_type': 'header'})
    rows.append(row('Increase/(Decrease) in Capital / Reserves', 'd_cap', indent=1))
    rows.append(row('Increase/(Decrease) in Loans', 'd_loan', indent=1))
    rows.append(row('= CASH FLOW FROM FINANCING ACTIVITY', 'cff', rtype='cff'))

    rows.append({'Particulars': '', '_type': 'spacer'})
    rows.append(row('NET CASH GENERATED / (USED)', 'net_cash', rtype='subtotal'))
    rows.append(row('Add: Opening Cash Balance', 'opening'))
    rows.append(row('= CLOSING CASH BALANCE (Computed)', 'closing_computed', rtype='recon'))
    rows.append(row('Actual Closing Cash Balance (from Balance Sheet)', 'closing_actual'))
    diff_row = row('Difference (should be ~0)', 'diff', rtype='recon')
    diff_total = sum(per_month.get(lbl, {}).get('diff', 0) for lbl in mo_labels)
    for lbl in mo_labels:
        tied = per_month.get(lbl, {}).get('tied', False)
        diff_row[lbl] = ("✅ " if tied else "⚠️ ") + diff_row[lbl]
    diff_row['Total'] = fmt_inr(diff_total) if diff_total else "—"
    rows.append(diff_row)

    return rows


def _interest_row(per_month, mo_labels):
    r = {'Particulars': '    Add/(Less): Interest Paid / (Received)', '_type': 'ledger'}
    total = 0.0
    for lbl in mo_labels:
        m = per_month.get(lbl, {})
        v = m.get('interest_paid', 0) - m.get('interest_received', 0)
        total += v
        r[lbl] = fmt_inr(v) if v else "—"
    r['Total'] = fmt_inr(total) if total else "—"
    return r


def _render_html_table(table_rows, mo_labels):
    STYLES = {
        'header':   'background:#1D3557;color:white;font-weight:700;font-size:0.85rem;',
        'ledger':   'background:white;color:#1f2937;font-size:0.82rem;',
        'subtotal': 'background:#e8f4f8;color:#1D3557;font-weight:600;font-size:0.83rem;',
        'cfo':      'background:#1D9E75;color:white;font-weight:700;font-size:0.88rem;',
        'cfi':      'background:#E07B39;color:white;font-weight:700;font-size:0.88rem;',
        'cff':      'background:#6C5CE7;color:white;font-weight:700;font-size:0.88rem;',
        'recon':    'background:#D85A30;color:white;font-weight:700;font-size:0.88rem;',
        'spacer':   'background:#f8f9fa;height:6px;',
    }

    all_cols = ["Particulars"] + mo_labels + ["Total"]
    th_style = "padding:6px 10px;text-align:right;white-space:nowrap;border:1px solid #dee2e6;"
    th_left  = "padding:6px 10px;text-align:left;white-space:nowrap;border:1px solid #dee2e6;"
    th_total = ("padding:6px 10px;text-align:right;white-space:nowrap;"
                "border:1px solid #dee2e6;border-left:2px solid #4a6fa5;background:#12233d;")

    html  = '<div style="overflow-x:auto;max-height:620px;overflow-y:auto;">'
    html += '<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:0.82rem;">'
    html += '<thead><tr style="background:#1D3557;color:white;position:sticky;top:0;">'
    for c in all_cols:
        if c == "Particulars":
            html += f'<th style="{th_left}">{c}</th>'
        elif c == "Total":
            html += f'<th style="{th_total}">{c}</th>'
        else:
            html += f'<th style="{th_style}">{c}</th>'
    html += '</tr></thead><tbody>'

    for r in table_rows:
        rtype  = r.get('_type', 'ledger')
        rstyle = STYLES.get(rtype, '')
        html  += f'<tr style="{rstyle}">'
        if rtype == 'spacer':
            html += f'<td colspan="{len(all_cols)}" style="padding:2px;"></td>'
        else:
            for c in all_cols:
                v     = r.get(c, '')
                align = 'left' if c == 'Particulars' else 'right'
                extra = 'border-left:2px solid #4a6fa5;font-weight:600;' if c == 'Total' else ''
                html += f'<td style="padding:5px 10px;text-align:{align};white-space:nowrap;border:1px solid #e5e7eb;{extra}">{v}</td>'
        html += '</tr>'

    html += '</tbody></table></div>'
    st.markdown(html, unsafe_allow_html=True)


# ── EXCEL EXPORT ──────────────────────────────────────────────────
def _generate_cashflow_excel(per_month, mo_labels, from_lbl, to_lbl, company_name):
    """Mirrors reports.py::_generate_excel() styling conventions exactly
    (same fonts, fills, borders, freeze panes, title/header approach)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    def fill(hex_):
        return PatternFill('solid', start_color=hex_, end_color=hex_)

    def border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    SECTION_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    SUBTOTAL_FONT = Font(name='Arial', bold=True, color='1D3557', size=9)
    LEDGER_FONT   = Font(name='Arial', size=9)
    NUM_FORMAT    = '#,##0.00'

    HEADER_FILL   = fill('1D3557')
    SECTION_FILL  = fill('2E75B6')
    SUBTOTAL_FILL = fill('D6E4F0')
    CFO_FILL      = fill('1D9E75')
    CFI_FILL      = fill('E07B39')
    CFF_FILL      = fill('6C5CE7')
    RECON_FILL    = fill('D85A30')
    LEDGER_FILL   = fill('FFFFFF')
    SPACER_FILL   = fill('F0F0F0')

    ws.merge_cells(f'A1:{get_column_letter(len(mo_labels)+2)}1')
    ws['A1'] = f"Cash Flow Statement (Indirect Method) — {company_name} | {from_lbl} to {to_lbl}"
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='1D3557')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Particulars'] + mo_labels + ['Total']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left' if ci == 1 else 'right', vertical='center')
        cell.border = border()
    ws.row_dimensions[2].height = 20

    ws.column_dimensions['A'].width = 42
    for ci in range(2, len(mo_labels) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = 'B3'

    row = [3]  # mutable box so nested fn can bump it

    def write_row(label, values_by_month, style, indent=0, is_diff=False):
        r = row[0]
        prefix = '    ' * indent
        c0 = ws.cell(row=r, column=1, value=prefix + label)
        c0.font, c0.fill = style['font'], style['fill']
        c0.alignment = Alignment(horizontal='left', vertical='center')
        c0.border = border()
        total = 0.0
        for ci, lbl in enumerate(mo_labels, 2):
            v = values_by_month.get(lbl, 0)
            total += v
            cell = ws.cell(row=r, column=ci, value=round(v, 2) if v else None)
            cell.font, cell.fill = style['font'], style['fill']
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border()
            if v:
                cell.number_format = NUM_FORMAT
        # Total column — simple sum of all months
        tcol = len(mo_labels) + 2
        tcell = ws.cell(row=r, column=tcol,
                        value=round(total, 2) if (total and values_by_month) else None)
        tcell.font, tcell.fill = style['font'], style['fill']
        tcell.alignment = Alignment(horizontal='right', vertical='center')
        tcell.border = border()
        if total and values_by_month:
            tcell.number_format = NUM_FORMAT
        ws.row_dimensions[r].height = 16
        row[0] += 1

    def vals(key, sign=1):
        return {lbl: per_month.get(lbl, {}).get(key, 0) * sign for lbl in mo_labels}

    def interest_vals():
        return {lbl: per_month.get(lbl, {}).get('interest_paid', 0)
                     - per_month.get(lbl, {}).get('interest_received', 0)
                for lbl in mo_labels}

    S_HEAD = {'font': SECTION_FONT, 'fill': SECTION_FILL}
    S_LED  = {'font': LEDGER_FONT, 'fill': LEDGER_FILL}
    S_SUB  = {'font': SUBTOTAL_FONT, 'fill': SUBTOTAL_FILL}
    S_CFO  = {'font': HEADER_FONT, 'fill': CFO_FILL}
    S_CFI  = {'font': HEADER_FONT, 'fill': CFI_FILL}
    S_CFF  = {'font': HEADER_FONT, 'fill': CFF_FILL}
    S_REC  = {'font': HEADER_FONT, 'fill': RECON_FILL}
    S_SPACE= {'font': LEDGER_FONT, 'fill': SPACER_FILL}

    write_row('I. CASH FLOW FROM OPERATING ACTIVITY (CFO)', {}, S_HEAD)
    write_row('Net Profit (as per P&L)', vals('net_profit'), S_LED, 1)
    write_row('Add: Depreciation', vals('depreciation'), S_LED, 1)
    write_row('Add/(Less): Interest Paid / (Received)', interest_vals(), S_LED, 1)
    write_row('= EBITDA / Operating Profit before Working Capital Changes',
              vals('ebitda'), S_SUB, 1)
    write_row('Working Capital Changes', {}, S_LED, 1)
    write_row('(Increase)/Decrease in Debtors', vals('d_debtors', -1), S_LED, 2)
    write_row('(Increase)/Decrease in Stock', vals('d_stock', -1), S_LED, 2)
    write_row('(Increase)/Decrease in Loans & Advances', vals('d_loanadv', -1), S_LED, 2)
    write_row('(Increase)/Decrease in Deposits', vals('d_deposits', -1), S_LED, 2)
    write_row('(Increase)/Decrease in Other Current Assets', vals('d_oca', -1), S_LED, 2)
    write_row('= Net (Decrease)/Increase in Current Assets', vals('wc_assets'), S_SUB, 2)
    write_row('Increase/(Decrease) in Creditors', vals('d_creditors'), S_LED, 2)
    write_row('Increase/(Decrease) in Provisions', vals('d_prov'), S_LED, 2)
    write_row('Increase/(Decrease) in Duties & Taxes', vals('d_duties'), S_LED, 2)
    write_row('Increase/(Decrease) in Other Current Liabilities', vals('d_ocl'), S_LED, 2)
    write_row('= Net Increase/(Decrease) in Current Liabilities', vals('wc_liab'), S_SUB, 2)
    write_row('= CASH FLOW FROM OPERATING ACTIVITY', vals('cfo'), S_CFO)

    write_row('', {}, S_SPACE)
    write_row('II. CASH FLOW FROM INVESTING ACTIVITY (CFI)', {}, S_HEAD)
    write_row('(Increase)/Decrease in Fixed Assets', vals('d_fa', -1), S_LED, 1)
    write_row('(Increase)/Decrease in Investments', vals('d_inv', -1), S_LED, 1)
    write_row('= CASH FLOW FROM INVESTING ACTIVITY', vals('cfi'), S_CFI)

    write_row('', {}, S_SPACE)
    write_row('III. CASH FLOW FROM FINANCING ACTIVITY (CFF)', {}, S_HEAD)
    write_row('Increase/(Decrease) in Capital / Reserves', vals('d_cap'), S_LED, 1)
    write_row('Increase/(Decrease) in Loans', vals('d_loan'), S_LED, 1)
    write_row('= CASH FLOW FROM FINANCING ACTIVITY', vals('cff'), S_CFF)

    write_row('', {}, S_SPACE)
    write_row('NET CASH GENERATED / (USED)', vals('net_cash'), S_SUB)
    write_row('Add: Opening Cash Balance', vals('opening'), S_LED)
    write_row('= CLOSING CASH BALANCE (Computed)', vals('closing_computed'), S_REC)
    write_row('Actual Closing Cash Balance (from Balance Sheet)', vals('closing_actual'), S_LED)
    write_row('Difference (should be ~0)', vals('diff'), S_REC)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
