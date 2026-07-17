"""
MIS Reports Page — P&L Detailed Report
Monthly breakup with all ledgers, Chart + Table format
"""
import streamlit as st
from datetime import date
from core.db import get_conn
from core.auth import is_admin
from core.theme import chart_layout, CHART_COLORS, CHART_PALETTE
from core.constants import MONTHS, TALLY_SECTION, SECTION_ORDER


def show_reports(user):
    conn = get_conn()

    # ── Read from shared sidebar filter (set by app.py) ──────
    company_id   = st.session_state.get('global_company_id')
    company_name = st.session_state.get('global_company_name', '')
    from_lbl     = st.session_state.get('global_from_lbl')
    to_lbl       = st.session_state.get('global_to_lbl')
    from_yr      = st.session_state.get('global_from_yr')
    from_mo      = st.session_state.get('global_from_mo')
    to_yr        = st.session_state.get('global_to_yr')
    to_mo        = st.session_state.get('global_to_mo')

    if not company_id or not from_lbl:
        conn.close()
        st.info("Select a company and date range from the sidebar to view reports.")
        return

    if (from_yr, from_mo) > (to_yr, to_mo):
        conn.close()
        st.error("'From' cannot be after 'To'.")
        return

    # Selected months list
    avail = conn.execute(
        "SELECT DISTINCT year, month FROM pl_data "
        "WHERE company_id=? ORDER BY year, month", (company_id,)
    ).fetchall()

    if not avail:
        conn.close()
        st.info("No data synced yet. Please sync from Tally first.")
        return

    sel_months = [
        (int(r['year']), int(r['month']))
        for r in avail
        if (int(r['year']), int(r['month'])) >= (from_yr, from_mo)
        and (int(r['year']), int(r['month'])) <= (to_yr, to_mo)
    ]
    mo_labels = [f"{MONTHS[m-1]}-{str(y)[2:]}" for y, m in sel_months]
    n_months  = len(sel_months)

    st.markdown(f"""
        <div class="page-sticky-header">
            <div class="page-title-row">📄 MIS Reports</div>
            <div class="page-company-row">🏢 {company_name}</div>
        </div>
    """, unsafe_allow_html=True)
    st.caption(f"📅 {from_lbl} → {to_lbl} | {n_months} months")
    st.markdown("---")

    # ── FETCH DATA ─────────────────────────────────────────
    pl_rows = conn.execute("""
        SELECT ledger_name, tally_group, mis_group, year, month, net
        FROM pl_data
        WHERE company_id=?
          AND ((year > ?) OR (year = ? AND month >= ?))
          AND ((year < ?) OR (year = ? AND month <= ?))
        ORDER BY tally_group, ledger_name, year, month
    """, (company_id, from_yr, from_yr, from_mo,
          to_yr, to_yr, to_mo)).fetchall()
    conn.close()

    if not pl_rows:
        st.info("No data for selected period.")
        return

    import pandas as pd
    import plotly.graph_objects as go
    import plotly.express as px

    # ── BUILD PIVOT TABLE ─────────────────────────────────
    data        = {}  # individual ledgers: {(sec,tg,ln): {(y,m): val}}
    group_totals= {}  # group totals from Tally group total rows

    for r in pl_rows:
        tg  = str(r['tally_group']  or '').lower().strip()
        mg  = str(r['mis_group']    or '').lower().strip()
        ln  = str(r['ledger_name']  or '').strip()
        yr  = int(r['year'])
        mo  = int(r['month'])
        val = r['net'] or 0
        lbl = f"{MONTHS[mo-1]}-{str(yr)[2:]}"

        # Capture Indirect Expenses group total BEFORE skip rules
        # Accumulate net (not abs) — take abs later for correct annual total
        if tg == 'indirect expenses' and ln.lower() == 'indirect expenses':
            group_totals.setdefault('overhead', {})
            group_totals['overhead'][lbl] = group_totals['overhead'].get(lbl, 0) + val
            continue

        # Capture Salaries and Bonus group total BEFORE skip rules
        if tg == 'salaries and bonus' and ln.lower() == 'salaries and bonus':
            group_totals.setdefault('overhead_sal', {})
            group_totals['overhead_sal'][lbl] = group_totals['overhead_sal'].get(lbl, 0) + val
            continue

        # Skip rows tagged as group totals
        if mg == '_group_total_':
            continue

        # Collect group total rows (ledger_name == tally_group)
        if ln.lower() == tg:
            if tg == 'sales accounts':
                group_totals.setdefault('revenue', {})
                group_totals['revenue'][lbl] = group_totals['revenue'].get(lbl,0) + val
            elif tg == 'direct incomes':
                group_totals.setdefault('dir_inc_group', {})
                group_totals['dir_inc_group'][lbl] = group_totals['dir_inc_group'].get(lbl,0) + val
            elif tg == 'cost of sales :':
                group_totals.setdefault('cogs', {})
                # Use net (not abs) — monthly COGS can be positive when closing > opening
                group_totals['cogs'][lbl] = group_totals['cogs'].get(lbl,0) + val
            continue

        sec = TALLY_SECTION.get(tg)
        if not sec:
            continue

        key = (sec, tg, ln)
        if key not in data:
            data[key] = {}
        data[key][(yr, mo)] = data[key].get((yr, mo), 0) + val

    # ── HELPER ────────────────────────────────────────────
    def fmt_cr(v):
        if v is None or v == 0: return "—"
        a = abs(v)
        if a >= 1e7:   return f"{'- ' if v<0 else ''}{a/1e7:.2f} Cr"
        elif a >= 1e5: return f"{'- ' if v<0 else ''}{a/1e5:.2f} L"
        else:          return f"{'- ' if v<0 else ''}₹{a:,.0f}"

    # ── BUILD SECTIONS (outside tabs — needed for Excel download) ──
    sections = {
        'revenue':    {},
        'dir_inc':    {},
        'opening':    {},
        'purchases':  {},
        'direct_exp': {},
        'closing':    {},
        'cos_net':    {},
        'ind_inc':    {},
        'overhead':   {},
    }
    for (sec, tg, ln), mo_data in data.items():
        if sec not in sections:
            continue
        if ln not in sections[sec]:
            sections[sec][ln] = {}
        for (yr, mo), val in mo_data.items():
            lbl = f"{MONTHS[mo-1]}-{str(yr)[2:]}"
            sections[sec][ln][lbl] = sections[sec][ln].get(lbl, 0) + val

    # ── REPORT TABS ────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Chart View",
        "📋 Detailed Table",
        "👤 Customer Ageing",
        "🏭 Vendor Ageing",
    ])

    # ── TAB 1: CHART VIEW ─────────────────────────────────
    with tab1:
        monthly = {lbl: {'revenue':0,'dir_inc':0,'cogs':0,'cos_net':0,
                         'opening':0,'purchases':0,'direct_exp':0,'closing':0,
                         'ind_inc':0,'overhead':0} for lbl in mo_labels}

        for (sec, tg, ln), mo_data in data.items():
            for (yr, mo), val in mo_data.items():
                lbl = f"{MONTHS[mo-1]}-{str(yr)[2:]}"
                if lbl in monthly and sec in monthly[lbl]:
                    monthly[lbl][sec] += abs(val)

        # Use group_totals as a FALLBACK only (not a forced override).
        # The sync parser doesn't emit a 'Sales Accounts' group-total
        # row anymore — only individual member ledgers — so any old
        # group-total row in the DB is a stale leftover that will
        # never refresh again. `monthly[lbl]['revenue']` computed just
        # above (summed from individual ledgers) is the live, correct
        # value; only fall back to group_totals if that sum came up
        # empty (e.g. genuinely no data for that month at all).
        for lbl in mo_labels:
            if 'revenue' in group_totals and monthly[lbl]['revenue'] == 0:
                monthly[lbl]['revenue'] = group_totals['revenue'].get(lbl, 0)
            if 'dir_inc_group' in group_totals:
                monthly[lbl]['dir_inc'] = max(
                    group_totals['dir_inc_group'].get(lbl, 0),
                    monthly[lbl]['dir_inc']
                )
            if 'cogs' in group_totals:
                monthly[lbl]['cogs'] = abs(group_totals['cogs'].get(lbl, 0))
            else:
                monthly[lbl]['cogs'] = (monthly[lbl]['opening'] +
                                        monthly[lbl]['purchases'] +
                                        monthly[lbl]['direct_exp'] -
                                        monthly[lbl]['closing'])
            if 'overhead' in group_totals or 'overhead_sal' in group_totals:
                ie_val  = abs(group_totals.get('overhead', {}).get(lbl, 0))
                sal_val = abs(group_totals.get('overhead_sal', {}).get(lbl, 0))
                monthly[lbl]['overhead'] = ie_val + sal_val

        labels   = mo_labels
        revenue  = [monthly[l]['revenue']  for l in labels]
        dir_inc  = [monthly[l]['dir_inc']  for l in labels]
        cogs     = [monthly[l]['cogs']     for l in labels]
        overhead = [monthly[l]['overhead'] for l in labels]
        ind_inc  = [monthly[l]['ind_inc']  for l in labels]
        gp       = [revenue[i] + dir_inc[i] - cogs[i] for i in range(len(labels))]
        np_      = [gp[i] + ind_inc[i] - overhead[i]  for i in range(len(labels))]

        col1, col2 = st.columns(2)

        with col1:
            st.markdown('<div class="chart-card"><p class="chart-title">📈 Revenue vs GP vs NP</p>',
                        unsafe_allow_html=True)
            fig = go.Figure()
            bar_colors = [CHART_PALETTE[i % len(CHART_PALETTE)] for i in range(len(labels))]
            fig.add_trace(go.Bar(
                x=labels, y=revenue, name='Revenue',
                marker=dict(color=bar_colors,
                            line=dict(color='rgba(11,15,26,0.25)', width=1))))
            fig.add_trace(go.Scatter(
                x=labels, y=revenue, mode='markers', name='',
                marker=dict(symbol='line-ew', size=18, line=dict(width=3, color='rgba(255,255,255,0.6)')),
                showlegend=False, hoverinfo='skip'))
            fig.add_trace(go.Scatter(x=labels, y=gp, name='Gross Profit',
                                     mode='lines+markers',
                                     line=dict(color=CHART_COLORS['blue'], width=2.5)))
            fig.add_trace(go.Scatter(x=labels, y=np_, name='Net Profit',
                                     mode='lines+markers',
                                     line=dict(color=CHART_COLORS['violet'], width=2.5, dash='dot')))
            chart_layout(fig, height=320,
                        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                    font=dict(color=CHART_COLORS['text'], size=10)))
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
            st.markdown('</div>', unsafe_allow_html=True)

        with col2:
            st.markdown('<div class="chart-card"><p class="chart-title">💸 COGS vs Overhead Trend</p>',
                        unsafe_allow_html=True)
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(x=labels, y=cogs, name='COGS',
                                  marker=dict(color=CHART_COLORS['amber'], opacity=0.9,
                                              line=dict(color='rgba(255,255,255,0.15)', width=1))))
            fig2.add_trace(go.Bar(x=labels, y=overhead, name='Overhead',
                                  marker=dict(color=CHART_COLORS['red'], opacity=0.9,
                                              line=dict(color='rgba(255,255,255,0.15)', width=1))))
            chart_layout(fig2, height=320, barmode='group',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                    font=dict(color=CHART_COLORS['text'], size=10)))
            st.plotly_chart(fig2, use_container_width=True, config={'displayModeBar': False})
            st.markdown('</div>', unsafe_allow_html=True)

        # GP% and NP% trend
        st.markdown('<div class="chart-card"><p class="chart-title">📊 GP% and NP% Monthly Trend</p>',
                    unsafe_allow_html=True)
        gp_pct = [round(gp[i]/revenue[i]*100,1) if revenue[i] else 0 for i in range(len(labels))]
        np_pct = [round(np_[i]/revenue[i]*100,1) if revenue[i] else 0 for i in range(len(labels))]
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=labels, y=gp_pct, name='GP%',
                                  mode='lines+markers+text',
                                  text=[f"{v}%" for v in gp_pct],
                                  textposition='top center',
                                  textfont=dict(color=CHART_COLORS['text']),
                                  line=dict(color=CHART_COLORS['mint'], width=2)))
        fig3.add_trace(go.Scatter(x=labels, y=np_pct, name='NP%',
                                  mode='lines+markers+text',
                                  text=[f"{v}%" for v in np_pct],
                                  textposition='bottom center',
                                  textfont=dict(color=CHART_COLORS['text']),
                                  line=dict(color=CHART_COLORS['violet'], width=2, dash='dot')))
        fig3.add_hline(y=0, line_dash='dash', line_color='rgba(255,255,255,0.25)', opacity=0.6)
        chart_layout(fig3, height=300,
                    legend=dict(orientation='h', font=dict(color=CHART_COLORS['text'], size=10)),
                    yaxis=dict(gridcolor=CHART_COLORS['grid'], color=CHART_COLORS['text'],
                                ticksuffix='%'))
        st.plotly_chart(fig3, use_container_width=True, config={'displayModeBar': False})
        st.markdown('</div>', unsafe_allow_html=True)

    # ── TAB 2: DETAILED TABLE ─────────────────────────────
    with tab2:

        # sections already built above (outside tabs)

        def section_total(sec_key, lbl):
            """Use group totals where available. Returns signed monthly values."""
            if sec_key == 'revenue' and 'revenue' in group_totals:
                return abs(group_totals['revenue'].get(lbl, 0))
            if sec_key == 'dir_inc':
                gt    = abs(group_totals.get('dir_inc_group', {}).get(lbl, 0))
                items = sum(abs(v.get(lbl, 0)) for v in sections['dir_inc'].values())
                return max(gt, items)
            if sec_key == 'cogs':
                if 'cogs' in group_totals:
                    # Return signed net — caller uses abs() for display
                    return group_totals['cogs'].get(lbl, 0)
                opening  = sum(abs(v.get(lbl,0)) for v in sections['opening'].values())
                purchase = sum(abs(v.get(lbl,0)) for v in sections['purchases'].values())
                direxp   = sum(abs(v.get(lbl,0)) for v in sections['direct_exp'].values())
                closing  = sum(abs(v.get(lbl,0)) for v in sections['closing'].values())
                return opening + purchase + direxp - closing
            if sec_key == 'overhead' and ('overhead' in group_totals or 'overhead_sal' in group_totals):
                ie_val  = abs(group_totals.get('overhead', {}).get(lbl, 0))
                sal_val = abs(group_totals.get('overhead_sal', {}).get(lbl, 0))
                return ie_val + sal_val
            return sum(abs(v.get(lbl, 0)) for v in sections[sec_key].values())

        def annual_total(sec_key):
            """Annual total using abs(sum(net)) — correct for sign-flip months."""
            if sec_key == 'revenue':
                return abs(sum(group_totals.get('revenue', {}).values()))
            if sec_key == 'dir_inc':
                return abs(sum(group_totals.get('dir_inc_group', {}).values()))
            if sec_key == 'cogs':
                return abs(sum(group_totals.get('cogs', {}).values()))
            if sec_key == 'overhead':
                return (abs(sum(group_totals.get('overhead', {}).values())) +
                        abs(sum(group_totals.get('overhead_sal', {}).values())))
            # Default: sum of monthly values
            return sum(section_total(sec_key, l) for l in mo_labels)

        # ── BUILD TABLE ROWS ──────────────────────────────
        table_rows = []

        def add_section(title, sec_key):
            sec_data = sections[sec_key]
            if not sec_data:
                return
            # Section header
            hrow = {'Particulars': title, '_type': 'header'}
            for lbl in mo_labels: hrow[lbl] = ''
            hrow['Total'] = ''
            table_rows.append(hrow)
            # Individual ledger rows
            for ln, mo_data in sorted(sec_data.items()):
                if all(abs(v) < 100 for v in mo_data.values()):
                    continue
                row = {'Particulars': f"  {ln}", '_type': 'ledger'}
                total = 0
                for lbl in mo_labels:
                    v = mo_data.get(lbl, 0)
                    row[lbl] = fmt_cr(abs(v)) if abs(v) >= 100 else '—'
                    total += abs(v)
                row['Total'] = fmt_cr(total)
                table_rows.append(row)
            # Subtotal
            srow = {'Particulars': '  Sub-Total', '_type': 'subtotal'}
            grand = 0
            for lbl in mo_labels:
                t = sum(abs(v.get(lbl, 0)) for v in sec_data.values())
                srow[lbl] = fmt_cr(t)
                grand += t
            srow['Total'] = fmt_cr(grand)
            table_rows.append(srow)

        add_section("📈 SALES ACCOUNTS (Revenue)",   'revenue')
        add_section("➕ DIRECT INCOMES",              'dir_inc')

        # COGS section — Tally style breakdown
        cogs_hrow = {'Particulars': '📦 COST OF GOODS SOLD', '_type': 'header'}
        for lbl in mo_labels: cogs_hrow[lbl] = ''
        cogs_hrow['Total'] = ''
        table_rows.append(cogs_hrow)

        def add_cogs_sub(title, sec_key, is_deduction=False):
            sec_data = sections.get(sec_key, {})
            if not sec_data: return
            # Sub-section header
            sh = {'Particulars': f'  {title}', '_type': 'subtotal'}
            for lbl in mo_labels: sh[lbl] = ''
            sh['Total'] = ''
            table_rows.append(sh)
            for ln, mo_data in sorted(sec_data.items()):
                if all(abs(v) < 100 for v in mo_data.values()): continue
                row = {'Particulars': f'    {ln}', '_type': 'ledger'}
                total = 0
                for lbl in mo_labels:
                    v = abs(mo_data.get(lbl, 0))
                    row[lbl] = fmt_cr(v) if v >= 100 else '—'
                    total += v
                row['Total'] = fmt_cr(total)
                table_rows.append(row)

        # Show breakdown if individual data available after re-sync
        # Otherwise show Cost of Sales net row
        has_breakdown = any(sections.get(k) for k in ('opening','purchases','direct_exp','closing'))
        if has_breakdown:
            add_cogs_sub('Opening Stock',       'opening')
            add_cogs_sub('Add: Purchase Accounts', 'purchases')
            add_cogs_sub('Add: Direct Expenses',   'direct_exp')
            add_cogs_sub('Less: Closing Stock',    'closing', is_deduction=True)
        else:
            crow = {'Particulars': '  Opening Stock + Purchases - Closing (Net)', '_type': 'ledger'}
            for lbl in mo_labels:
                crow[lbl] = fmt_cr(section_total('cogs', lbl))
            crow['Total'] = fmt_cr(sum(section_total('cogs', l) for l in mo_labels))
            table_rows.append(crow)
            # Direct Expenses
            for ln, mo_data in sorted(sections.get('direct_exp', {}).items()):
                if all(abs(v) < 100 for v in mo_data.values()): continue
                drow = {'Particulars': f'  {ln}', '_type': 'ledger'}
                dtotal = 0
                for lbl in mo_labels:
                    v = abs(mo_data.get(lbl, 0))
                    drow[lbl] = fmt_cr(v) if v >= 100 else '—'
                    dtotal += v
                drow['Total'] = fmt_cr(dtotal)
                table_rows.append(drow)

        # COGS subtotal
        csrow = {'Particulars': '  Sub-Total', '_type': 'subtotal'}
        for lbl in mo_labels:
            v = section_total('cogs', lbl)
            csrow[lbl] = fmt_cr(abs(v))
        csrow['Total'] = fmt_cr(annual_total('cogs'))
        table_rows.append(csrow)

        # GP row
        table_rows.append({'Particulars': '', '_type': 'spacer'})
        grow = {'Particulars': '💹 GROSS PROFIT', '_type': 'gp'}
        for lbl in mo_labels:
            g = (section_total('revenue', lbl) +
                 section_total('dir_inc', lbl) -
                 abs(section_total('cogs', lbl)))
            grow[lbl] = fmt_cr(g)
        gp_annual = (annual_total('revenue') + annual_total('dir_inc') - annual_total('cogs'))
        grow['Total'] = fmt_cr(gp_annual)
        table_rows.append(grow)
        table_rows.append({'Particulars': '', '_type': 'spacer'})

        add_section("💰 INDIRECT INCOMES",            'ind_inc')
        add_section("💸 INDIRECT EXPENSES (Overhead)",'overhead')

        # NP row
        table_rows.append({'Particulars': '', '_type': 'spacer'})
        nrow = {'Particulars': '🏆 NET PROFIT', '_type': 'np'}
        for lbl in mo_labels:
            n = (section_total('revenue', lbl) +
                 section_total('dir_inc', lbl) -
                 abs(section_total('cogs', lbl)) +
                 section_total('ind_inc', lbl) -
                 section_total('overhead', lbl))
            nrow[lbl] = fmt_cr(n)
        ii_annual = annual_total('ind_inc')
        oh_annual = annual_total('overhead')
        np_annual = gp_annual + ii_annual - oh_annual
        nrow['Total'] = fmt_cr(np_annual)
        table_rows.append(nrow)

        # ── DOWNLOAD — dropdown: .xlsx or .csv, filename includes client ──
        report_base_name = f"P&L Detailed — {company_name}".strip()
        # Sanitize for filesystem (Windows-illegal chars)
        for ch in '\\/:*?"<>|':
            report_base_name = report_base_name.replace(ch, '-')

        col_title, col_dl = st.columns([6, 1])
        with col_title:
            st.markdown("**📋 P&L Detailed — Monthly Breakup**")
        with col_dl:
            with st.popover("📥 Export", use_container_width=True):
                st.caption(report_base_name)
                excel_bytes = _generate_excel(
                    sections, group_totals, mo_labels, sel_months, from_lbl, to_lbl
                )
                st.download_button(
                    label="⬇️ Download as .xlsx",
                    data=excel_bytes,
                    file_name=f"{report_base_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_xlsx",
                )
                csv_df  = pd.DataFrame(table_rows)
                csv_cols = ["Particulars"] + mo_labels + ["Total"]
                csv_df  = csv_df.reindex(columns=csv_cols, fill_value="")
                csv_bytes = csv_df.to_csv(index=False).encode("utf-8-sig")
                st.download_button(
                    label="⬇️ Download as .csv",
                    data=csv_bytes,
                    file_name=f"{report_base_name}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="dl_csv",
                )

        # ── RENDER HTML TABLE ─────────────────────────────
        STYLES = {
            'header':   'background:#1D3557;color:white;font-weight:700;font-size:0.85rem;',
            'ledger':   'background:white;color:#1f2937;font-size:0.82rem;',
            'subtotal': 'background:#e8f4f8;color:#1D3557;font-weight:600;font-size:0.83rem;',
            'gp':       'background:#1D9E75;color:white;font-weight:700;font-size:0.88rem;',
            'np':       'background:#D85A30;color:white;font-weight:700;font-size:0.88rem;',
            'spacer':   'background:#f8f9fa;height:6px;',
        }

        all_cols = ["Particulars"] + mo_labels + ["Total"]
        th_style = "padding:6px 10px;text-align:right;white-space:nowrap;border:1px solid #dee2e6;"
        th_left  = "padding:6px 10px;text-align:left;white-space:nowrap;border:1px solid #dee2e6;"

        html  = '<div style="overflow-x:auto;max-height:620px;overflow-y:auto;">'
        html += '<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:0.82rem;">'
        html += '<thead><tr style="background:#1D3557;color:white;position:sticky;top:0;">'
        for c in all_cols:
            html += f'<th style="{th_left if c=="Particulars" else th_style}">{c}</th>'
        html += '</tr></thead><tbody>'

        for row in table_rows:
            rtype  = row.get('_type', 'ledger')
            rstyle = STYLES.get(rtype, '')
            html  += f'<tr style="{rstyle}">'
            if rtype == 'spacer':
                html += f'<td colspan="{len(all_cols)}" style="padding:2px;"></td>'
            else:
                for c in all_cols:
                    v     = row.get(c, '')
                    align = 'left' if c == 'Particulars' else 'right'
                    html += f'<td style="padding:5px 10px;text-align:{align};white-space:nowrap;border:1px solid #e5e7eb;">{v}</td>'
            html += '</tr>'

        html += '</tbody></table></div>'
        st.markdown(html, unsafe_allow_html=True)

    # ── TAB 3: CUSTOMER AGEING ────────────────────────────
    with tab3:
        _show_ageing_tab(company_id, 'customer', 'Customer Ageing', '👤')

    # ── TAB 4: VENDOR AGEING ──────────────────────────────
    with tab4:
        _show_ageing_tab(company_id, 'vendor', 'Vendor Ageing', '🏭')


def _show_ageing_tab(company_id, party_type, label, icon):
    """Render ageing report tab inside reports page"""
    import sqlite3, pandas as pd
    import plotly.graph_objects as go

    DB = "data/mis_portal.db"

    BUCKETS = [
        (0,   30,  "0-30 Days"),
        (31,  60,  "31-60 Days"),
        (61,  90,  "61-90 Days"),
        (91,  180, "91-180 Days"),
        (181, 365, "181-365 Days"),
        (366, None,"1 Year+"),
    ]
    BUCKET_COLS = [b[2] for b in BUCKETS]

    def fmt(v):
        if not v or v == 0: return "—"
        a = abs(v)
        if a >= 1e7:    return f"₹{a/1e7:.2f} Cr"
        elif a >= 1e5:  return f"₹{a/1e5:.2f} L"
        elif a >= 1000: return f"₹{a/1000:.1f}K"
        return f"₹{a:,.0f}"

    def bucket(days):
        for lo, hi, lbl in BUCKETS:
            if hi is None:
                if days >= lo: return lbl
            elif lo <= days <= hi: return lbl
        return "1 Year+"

    # Load data
    conn = sqlite3.connect(DB)
    rows = conn.execute("""
        SELECT party_name, bill_ref, bill_date, amount, days_overdue
        FROM ageing_data WHERE company_id=? AND party_type=?
        ORDER BY party_name, days_overdue DESC
    """, (company_id, party_type)).fetchall()
    last_sync = conn.execute(
        "SELECT MAX(synced_at) FROM ageing_data WHERE company_id=? AND party_type=?",
        (company_id, party_type)
    ).fetchone()[0]
    conn.close()

    if not rows:
        st.info(f"No {label} data found. Run sync from Tally first.")
        st.code("python sync_ageing.py", language="bash")
        return

    # Build party-wise buckets
    party_data = {}
    for r in rows:
        nm = r[0]; amt = r[3]; days = r[4]
        bkt = bucket(days)
        if nm not in party_data:
            party_data[nm] = {'Total':0, **{b:0 for b in BUCKET_COLS}}
        party_data[nm][bkt]   += amt
        party_data[nm]['Total'] += amt

    df = pd.DataFrame.from_dict(party_data, orient='index').reset_index()
    df.rename(columns={'index':'Party Name'}, inplace=True)
    df = df.sort_values('Total', ascending=False).reset_index(drop=True)

    # Totals row
    totals = {'Party Name':'📊 TOTAL'}
    for col in BUCKET_COLS + ['Total']:
        totals[col] = df[col].sum()
    df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

    total_os   = df[df['Party Name']=='📊 TOTAL']['Total'].values[0]
    total_curr = df[df['Party Name']=='📊 TOTAL']['0-30 Days'].values[0]
    total_over = total_os - total_curr
    n_parties  = len(df) - 1

    # KPIs
    c1,c2,c3,c4 = st.columns(4)
    c1.metric(f"{icon} Total Outstanding", fmt(total_os))
    c2.metric("✅ Current (0-30 Days)",    fmt(total_curr))
    c3.metric("⚠️ Overdue (31+ Days)",    fmt(total_over))
    c4.metric("👥 Parties",               str(n_parties))

    # Bar chart
    total_row    = df[df['Party Name']=='📊 TOTAL'].iloc[0]
    bucket_vals  = {b: total_row[b] for b in BUCKET_COLS if total_row[b] > 0}
    if bucket_vals:
        colors = ['#00C896','#00A3FF','#FFB347','#FF7F7F','#CF6FC4','#FF4444']
        fig = go.Figure(go.Bar(
            x=list(bucket_vals.keys()),
            y=[v/1e7 for v in bucket_vals.values()],
            marker=dict(color=colors[:len(bucket_vals)],
                        line=dict(color='rgba(11,15,26,0.15)', width=1)),
            text=[fmt(v) for v in bucket_vals.values()],
            textposition='outside',
            textfont=dict(color=CHART_COLORS['text_hi']),
        ))
        chart_layout(fig, yaxis_title="Amount (Cr)", height=280,
                     margin=dict(t=20, b=10, l=40, r=20))
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

    # Table
    st.markdown(f"**{icon} {label} — Party-wise Breakup**")
    display = df[['Party Name'] + BUCKET_COLS + ['Total']].copy()
    for col in BUCKET_COLS + ['Total']:
        display[col] = display[col].apply(lambda v: fmt(v) if isinstance(v,(int,float)) else v)

    st.dataframe(display, use_container_width=True,
                 height=min(550, (len(display)+1)*35+38), hide_index=True)

    # Drill-down
    with st.expander("🔍 Bill-level Detail by Party"):
        parties = sorted(set(r[0] for r in rows))
        sel = st.selectbox("Select Party", parties, key=f"ag_party_{party_type}")
        bills = [r for r in rows if r[0] == sel]
        if bills:
            bd = pd.DataFrame([{
                'Bill Date': r[2] or '—',
                'Bill Ref':  r[1] or '—',
                'Amount':    fmt(r[3]),
                'Days Overdue': r[4],
                'Bucket':    bucket(r[4]),
            } for r in sorted(bills, key=lambda x: x[4], reverse=True)])
            st.dataframe(bd, use_container_width=True, hide_index=True)
            st.caption(f"Total: **{fmt(sum(r[3] for r in bills))}** | {len(bills)} bills")

    if last_sync:
        st.caption(f"Last synced: {last_sync}")


def _generate_excel(sections, group_totals, mo_labels, sel_months, from_lbl, to_lbl):
    """Generate formatted P&L Excel — values match Dashboard exactly"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Detailed"

    # ── STYLES ─────────────────────────────────────────────
    def fill(hex_):
        return PatternFill('solid', start_color=hex_, end_color=hex_)

    def border():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    SECTION_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    SUBTOTAL_FONT = Font(name='Arial', bold=True, color='1D3557', size=9)
    LEDGER_FONT   = Font(name='Arial', size=9)
    GP_FONT       = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    NP_FONT       = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    NUM_FORMAT    = '#,##0.00'

    HEADER_FILL   = fill('1D3557')
    SECTION_FILL  = fill('2E75B6')
    SUBTOTAL_FILL = fill('D6E4F0')
    GP_FILL       = fill('1D9E75')
    NP_FILL       = fill('D85A30')
    LEDGER_FILL   = fill('FFFFFF')
    ALT_FILL      = fill('F8FAFB')
    SPACER_FILL   = fill('F0F0F0')

    # ── TITLE ──────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(len(mo_labels)+2)}1')
    ws['A1'] = f"P&L Detailed Report  |  {from_lbl} to {to_lbl}"
    ws['A1'].font    = Font(name='Arial', bold=True, size=13, color='1D3557')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── COLUMN HEADERS ─────────────────────────────────────
    headers = ['Particulars'] + mo_labels + ['Total']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(
            horizontal='left' if ci == 1 else 'right',
            vertical='center'
        )
        cell.border = border()
    ws.row_dimensions[2].height = 20

    # Column widths
    ws.column_dimensions['A'].width = 38
    for ci in range(2, len(mo_labels) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Freeze panes
    ws.freeze_panes = 'B3'

    row = 3

    def write_row(label, values, style, indent=0):
        nonlocal row
        prefix = '    ' * indent
        ws.cell(row=row, column=1, value=prefix + label).font = style['font']
        ws.cell(row=row, column=1).fill      = style['fill']
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=1).border    = border()

        for ci, v in enumerate(values, 2):
            cell = ws.cell(row=row, column=ci, value=v if v else None)
            cell.font      = style['font']
            cell.fill      = style['fill']
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border    = border()
            if v and isinstance(v, (int, float)) and v != 0:
                cell.number_format = NUM_FORMAT
        ws.row_dimensions[row].height = 16
        row += 1

    def section_total(sec_key, lbl):
        if sec_key == 'revenue' and 'revenue' in group_totals:
            return abs(group_totals['revenue'].get(lbl, 0))
        if sec_key == 'dir_inc':
            gt    = abs(group_totals.get('dir_inc_group', {}).get(lbl, 0))
            items = sum(abs(v.get(lbl,0)) for v in sections.get('dir_inc',{}).values())
            return max(gt, items)
        if sec_key == 'cogs':
            if 'cogs' in group_totals:
                return abs(group_totals['cogs'].get(lbl, 0))
            o = sum(abs(v.get(lbl,0)) for v in sections.get('opening',{}).values())
            p = sum(abs(v.get(lbl,0)) for v in sections.get('purchases',{}).values())
            d = sum(abs(v.get(lbl,0)) for v in sections.get('direct_exp',{}).values())
            c = sum(abs(v.get(lbl,0)) for v in sections.get('closing',{}).values())
            return o + p + d - c
        if sec_key == 'overhead' and ('overhead' in group_totals or 'overhead_sal' in group_totals):
            ie_val  = abs(group_totals.get('overhead', {}).get(lbl, 0))
            sal_val = abs(group_totals.get('overhead_sal', {}).get(lbl, 0))
            return ie_val + sal_val
        return sum(abs(v.get(lbl,0)) for v in sections.get(sec_key,{}).values())

    def add_section(title, sec_key):
        nonlocal row
        sec_data = sections.get(sec_key, {})
        if not sec_data: return
        write_row(title, [''] * (len(mo_labels) + 1),
                  {'font': SECTION_FONT, 'fill': SECTION_FILL})
        alt = False
        for ln, mo_data in sorted(sec_data.items()):
            if all(abs(v) < 100 for v in mo_data.values()): continue
            vals = [round(abs(mo_data.get(l,0))/1e7,4) if abs(mo_data.get(l,0))>=100
                    else None for l in mo_labels]
            total = sum(abs(mo_data.get(l,0)) for l in mo_labels)
            vals.append(round(total/1e7,4) if total>=100 else None)
            f = ALT_FILL if alt else LEDGER_FILL
            write_row(ln, vals, {'font': LEDGER_FONT, 'fill': f}, indent=1)
            alt = not alt
        sub_vals = [round(section_total(sec_key,l)/1e7,4) for l in mo_labels]
        sub_vals.append(round(sum(section_total(sec_key,l) for l in mo_labels)/1e7,4))
        write_row('Sub-Total', sub_vals, {'font': SUBTOTAL_FONT, 'fill': SUBTOTAL_FILL})

    def add_cogs_sub(title, sec_key, indent=1):
        nonlocal row
        sec_data = sections.get(sec_key, {})
        if not sec_data: return
        write_row(title, [''] * (len(mo_labels) + 1),
                  {'font': SUBTOTAL_FONT, 'fill': SUBTOTAL_FILL}, indent=indent)
        alt = False
        for ln, mo_data in sorted(sec_data.items()):
            if all(abs(v) < 100 for v in mo_data.values()): continue
            vals = [round(abs(mo_data.get(l,0))/1e7,4) if abs(mo_data.get(l,0))>=100
                    else None for l in mo_labels]
            total = sum(abs(mo_data.get(l,0)) for l in mo_labels)
            vals.append(round(total/1e7,4) if total>=100 else None)
            f = ALT_FILL if alt else LEDGER_FILL
            write_row(ln, vals, {'font': LEDGER_FONT, 'fill': f}, indent=indent+1)
            alt = not alt

    # ── ADD SECTIONS ───────────────────────────────────────
    add_section('SALES ACCOUNTS (Revenue)', 'revenue')
    add_section('DIRECT INCOMES',           'dir_inc')

    # ── COGS SECTION ───────────────────────────────────────
    write_row('COST OF GOODS SOLD',
              [''] * (len(mo_labels) + 1),
              {'font': SECTION_FONT, 'fill': SECTION_FILL})

    has_breakdown = any(sections.get(k) for k in ('opening','purchases','direct_exp','closing'))
    if has_breakdown:
        add_cogs_sub('Opening Stock',          'opening')
        add_cogs_sub('Add: Purchase Accounts', 'purchases')
        add_cogs_sub('Add: Direct Expenses',   'direct_exp')
        add_cogs_sub('Less: Closing Stock',    'closing')
    else:
        cos_vals = [round(section_total('cogs',l)/1e7,4) for l in mo_labels]
        cos_vals.append(round(sum(cos_vals),4))
        write_row('  Opening Stock + Purchases - Closing Stock (Net)',
                  cos_vals, {'font': LEDGER_FONT, 'fill': LEDGER_FILL})
        alt = False
        for ln, mo_data in sorted(sections.get('direct_exp',{}).items()):
            if all(abs(v) < 100 for v in mo_data.values()): continue
            vals = [round(abs(mo_data.get(l,0))/1e7,4) if abs(mo_data.get(l,0))>=100
                    else None for l in mo_labels]
            total = sum(abs(mo_data.get(l,0)) for l in mo_labels)
            vals.append(round(total/1e7,4) if total>=100 else None)
            f = ALT_FILL if alt else LEDGER_FILL
            write_row(f'  {ln}', vals, {'font': LEDGER_FONT, 'fill': f})
            alt = not alt

    # COGS subtotal — monthly values + correct annual total
    cos_vals = [round(section_total('cogs',l)/1e7,4) for l in mo_labels]
    # Annual total: abs(sum of monthly net) — avoids sign-flip inflation
    if 'cogs' in group_totals:
        annual_cogs = abs(sum(group_totals['cogs'].values())) / 1e7
    else:
        annual_cogs = sum(cos_vals)
    cos_vals.append(round(annual_cogs, 4))
    write_row('Sub-Total', cos_vals, {'font': SUBTOTAL_FONT, 'fill': SUBTOTAL_FILL})

    # Spacer
    for ci in range(1, len(mo_labels) + 3):
        ws.cell(row=row, column=ci).fill = SPACER_FILL
    ws.row_dimensions[row].height = 6
    row += 1

    # GP Row — monthly + correct annual total
    gp_vals = []
    for lbl in mo_labels:
        r_v = section_total('revenue', lbl)
        d_v = section_total('dir_inc', lbl)
        c_v = section_total('cogs', lbl)
        gp_vals.append(round((r_v + d_v - c_v)/1e7, 4))
    # Annual GP = abs(rev_net) + abs(dir_inc_net) - abs(cogs_net)
    annual_rev  = abs(sum(group_totals.get('revenue',{}).values()))
    annual_di   = abs(sum(group_totals.get('dir_inc_group',{}).values()))
    annual_cogs2= abs(sum(group_totals.get('cogs',{}).values()))
    annual_gp   = (annual_rev + annual_di - annual_cogs2) / 1e7
    gp_vals.append(round(annual_gp, 4))
    write_row('GROSS PROFIT (c/o)', gp_vals, {'font': GP_FONT, 'fill': GP_FILL})

    # Spacer
    for ci in range(1, len(mo_labels) + 3):
        ws.cell(row=row, column=ci).fill = SPACER_FILL
    ws.row_dimensions[row].height = 6
    row += 1

    add_section('INDIRECT INCOMES',            'ind_inc')
    add_section('INDIRECT EXPENSES (Overhead)', 'overhead')

    # Spacer
    for ci in range(1, len(mo_labels) + 3):
        ws.cell(row=row, column=ci).fill = SPACER_FILL
    ws.row_dimensions[row].height = 6
    row += 1

    # NP Row — monthly + correct annual total
    np_vals = []
    for lbl in mo_labels:
        r_v = section_total('revenue', lbl)
        d_v = section_total('dir_inc', lbl)
        c_v = section_total('cogs', lbl)
        i_v = section_total('ind_inc', lbl)
        o_v = section_total('overhead', lbl)
        np_vals.append(round(((r_v+d_v-c_v) + i_v - o_v)/1e7, 4))
    # Annual NP = Annual GP + Annual IndInc - Annual Overhead
    annual_ii_total = sum(
        abs(v) for mo_d in sections.get('ind_inc',{}).values()
        for v in mo_d.values()
    ) / 1e7
    _annual_oh = (abs(sum(group_totals.get('overhead',{}).values())) +
                  abs(sum(group_totals.get('overhead_sal',{}).values()))) / 1e7
    annual_np = annual_gp + annual_ii_total - _annual_oh
    np_vals.append(round(annual_np, 4))
    write_row('NET PROFIT (Nett)', np_vals, {'font': NP_FONT, 'fill': NP_FILL})

    # ── NOTE ROW ───────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1,
            value="Note: All values in Crores (Cr). Generated by MIS Portal.")
    ws.cell(row=row, column=1).font = Font(name='Arial', size=8,
                                           italic=True, color='888888')

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
